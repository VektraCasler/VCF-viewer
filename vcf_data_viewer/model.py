# vcf_data_viewer/model.py
''' This holds the data model for the vcf-data-viewer application. '''

# IMPORTS ------------------------------------------------

from .global_variables import *
from .infotrack_db import *
from .bed_db import *
import openpyxl

# VARIABLES ----------------------------------------------

ADDON_LIST = [
    'amp_ID', 
    'Cytoband', 
    'MANE_transcript (GRCh38)', 
    'Genexus_transcript (GRCh37)', 
    'Genexus_Exon(s)', 
    'Genexus_codons',
    'tier',
    'test_tissue',
]

# CLASSES ------------------------------------------------

class DataModel():

    def __init__(self) -> None:

        self.variables = dict()
        self.filename = str()
        self.selected_disposistion = "None"
        self.selection_index = 0

        self.bdb = BedLookupTable()
        self.idb = InfotrackLookupTable()
        
        self.DATA_FIELDS = list()

        return None

    def determine_data_fields(self, excel_columns) -> None:
        """ Method to determine if this is an already annotated excel file, or if it is an original one.  Sets the data field (columns) list appropriately. """

        return None

    def load_file(self) -> None:
        """ Recieves the filename from the view, attempts to load. """

        workbook = openpyxl.load_workbook(self.filename, data_only=True, read_only=True)
        self.variant_list = list()

        # Read in the columns names and make a list, then check the length to see if it's been opened before.
        self.column_list = [cell.value for cell in workbook.worksheets[0][1]]
        if len(self.column_list) < 70:
            self.new_file = True
        else:
            self.new_file = False

        # capturing the workbook sheet names as dispositions 
        for disposition in workbook.sheetnames:

            # Grab the appropriate worksheet and make it active
            sheet = workbook[disposition]

            for row in sheet.iter_rows(min_row=2):

                # prep a dictionary to hold row data
                row_dict = dict()

                # Step through all VCF fields
                for x in range(len(self.column_list)):
                    # print(x, row[x].value)
                    try:
                        row_dict[self.column_list[x]] = row[x].value
                    except:
                        row_dict[self.column_list[x]] = None


                # making some temp variables for clarity during coding
                gene = row_dict['Variant Annotation: Gene']
                bp = int(row_dict['Original Input: Pos'])
                c_dot = row_dict['Variant Annotation: cDNA change']

                # self.column_list = [*self.column_list, *ADDON_LIST]

                if self.new_file:
                    # ANNOTATING FROM THE BED FILE -----------------------------------------------
                    for item in ADDON_LIST[:-2]:
                        row_dict[item] = self.bdb.get_data(gene, bp, item)
                    # ANNOTATING FROM THE INFOTRACK DB FILE --------------------------------------
                    row_dict['tier'] = self.idb.recommend_tier(gene, c_dot, 'Blood') # Genexys is just blood samples currently
                    row_dict['test_tissue'] = self.idb.get_tissue_list(gene, c_dot)

                else:
                    extended_range = (len(self.column_list), len(self.column_list)+len(ADDON_LIST))
                    for x in range(extended_range[0], extended_range[1]):
                        # print(x, row[x].value)
                        try:
                            row_dict[self.column_list[x]] = row[x].value
                        except:
                            row_dict[self.column_list[x]] = None

                # ignores any sheet names that "non-standard" dispositions.  
                # The first sheet in the excel file typically has such a name.
                if disposition in DISPOSITIONS:
                    row_dict['Disposition'] = disposition
                else:
                    row_dict['Disposition'] = "None"

                self.variant_list.append(row_dict.copy())

        return None

    def save_file(self) -> None:
        """ Writes the sorted data out to disk with marker on the filename to denote file has been processed. """

        # appending a "(sorted)" note to the filename, but checking if that note is already there
        if VCF_FILE_SETTINGS['filename_addon'] in self.filename:
            filename = self.filename
        else:
            filename = self.filename[:-(len(VCF_FILE_SETTINGS['excel_extension']))] + VCF_FILE_SETTINGS['filename_addon'] + VCF_FILE_SETTINGS['excel_extension']

        # Openpyxl package work
        wb = openpyxl.Workbook()
        for tab in DISPOSITIONS:
            wb.create_sheet(tab)
            wb.active = wb[tab]
            ws = wb.active
            row = [column for column in self.column_list]
            ws.append(row)
            for entry in self.variant_list:
                if entry['Disposition'] == tab:
                    row = [entry[x] for x in self.column_list]
                    ws.append(row)

        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) # removing the default "Sheet" from openpyxl
        wb.save(filename)

        return None
    
    def change_disposition(self, selection: str, disposition: str, update_dict: dict) -> None:
        """ Updates the disposition of the selected record. """

        # Updating all the fields in the model instance of the record
        for field in self.column_list:
            self.variant_list[selection][field] = update_dict[field]
        self.variant_list[selection]['Disposition'] = disposition

        return None

    def count_dispositions(self, disposition) -> int:
        """ Method to count the number of records in the variant list with the passed disposition. """

        counter = 0

        for row in self.variant_list:
            if disposition == row['Disposition']:
                counter += 1

        return counter

    def output_text_files(self) -> None:
        """ Placeholder method which will eventually output the needed text files for the reporting script. """

        pass

        return None

# MAIN LOOP ----------------------------------------------

def main() -> None:

    model = DataModel()

    return None

if __name__ == '__main__':
    main()