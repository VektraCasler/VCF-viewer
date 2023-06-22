# excel_test.py
''' Practice with opening and importing excel sheets '''

# IMPORTS ------------------------------------------------

import openpyxl
import pandas as pd

# VARIABLES ----------------------------------------------

FILENAME = 'test_data/50-22.final.report.xlsx'
VCF_FIELDS = [
    "Original Input: Chrom",
    "Original Input: Pos",
    "Original Input: Reference allele",
    "Original Input: Alternate allele",
    "Variant Annotation: Gene",
    "Variant Annotation: cDNA change",
    "Variant Annotation: Protein Change",
    "Variant Annotation: RefSeq",
    "VCF: AF",
    "VCF: FAO",
    "VCF: FDP",
    "VCF: HRUN",
    "VCF: Filter",
    "VCF: Genotype",
    "COSMIC: ID",
    "COSMIC: Variant Count",
    "COSMIC: Variant Count (Tissue)",
    "ClinVar: ClinVar ID",
    "ClinVar: Clinical Significance",
    "gnomAD3: Global AF",
    "PhyloP: Vert Score",
    "CADD: Phred",
    "PolyPhen-2: HDIV Prediction",
    "SIFT: Prediction",
    "VCF: FSAF",
    "VCF: FSAR",
    "VCF: FSRF",
    "VCF: FSRR",
    "VCF: Fisher Odds Ratio",
    "VCF: Fisher P Value",
    "VCF: Binom Proportion",
    "VCF: Binom P Value",
    "Mpileup Qual: Read Depth",
    "Mpileup Qual: Start Reads",
    "Mpileup Qual: Stop Reads",
    "Mpileup Qual: Filtered Reference Forward Read Depth",
    "Mpileup Qual: Filtered Reference Reverse Read Depth",
    "Mpileup Qual: Unfiltered Reference Forward Read Depth",
    "Mpileup Qual: Unfiltered Reference Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Forward Read Depth",
    "Mpileup Qual: Filtered Variant Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Binomial Proportion",
    "Mpileup Qual: Filtered Variant Binomial P Value",
    "Mpileup Qual: Filtered Variant Fishers Odds Ratio",
    "Mpileup Qual: Filtered Variant Fishers P Value",
    "Mpileup Qual: Filtered VAF",
    "Mpileup Qual: Unfiltered Variant Forward Read Depth",
    "Mpileup Qual: Unfiltered Variant Reverse Read Depth",
    "Mpileup Qual: Unfiltered Variant Binomial Proportion",
    "Mpileup Qual: Unfiltered Variant Binomial P Value",
    "Mpileup Qual: Unfiltered Variant Fishers Odds Ratio",
    "Mpileup Qual: Unfiltered Variant Fishers P Value",
    "Mpileup Qual: Unfiltered VAF",
    "VCF: LEN",
    "VCF: QD",
    "VCF: STB",
    "VCF: STBP",
    "VCF: SVTYPE",
    "VCF: TYPE",
    "VCF: QUAL",
    "Variant Annotation: Coding",
    "Variant Annotation: Sequence Ontology",
    "Variant Annotation: Transcript",
    "Variant Annotation: All Mappings",
    "UniProt (GENE): Accession Number",
    "dbSNP: rsID",
    "MDL: Sample Count",
    "MDL: Variant Frequency",
    "MDL: Sample List",
]

# MAIN LOOP ----------------------------------------------

def read_excel_file_to_dictionary(filename:str) -> dict:
    """ Loads an Excel worksheet, then reads all sheets for variant information. """
    workbook = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    variant_list = list()
    for disposition in workbook.sheetnames:
        sheet = workbook[disposition]
        for row in sheet.iter_rows(min_row=2):
            row_dict = dict()
            for x in range(len(VCF_FIELDS)):
                row_dict[VCF_FIELDS[x]] = row[x].value
            row_dict['Disposition'] = disposition
            variant_list.append(row_dict.copy())
    return(variant_list)

def main():
    variant_list = read_excel_file_to_dictionary(FILENAME)
    DF = pd.DataFrame(variant_list)
    print(DF.head(10))
    return

if __name__ == '__main__':
    main()