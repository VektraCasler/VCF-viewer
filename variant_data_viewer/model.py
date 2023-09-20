# variant_data_viewer/model.py
""" This holds the data model for the variant-data-viewer application. """

# IMPORTS ------------------------------------------------

from tkinter import messagebox as mb
from .global_variables import *
from .infotrack_db import *
from .bed_db import *
import openpyxl

# VARIABLES ----------------------------------------------

ADDON_LIST = [
    "amp_ID",
    "Cytoband",
    "MANE_transcript (GRCh38)",
    "Genexus_transcript (GRCh37)",
    "Genexus_Exon(s)",
    "Genexus_codons",
    "tier",
    "test_tissue",
]
DATA_FIELDS = [
    "Original Input: Chrom",
    "Original Input: Pos",
    "Original Input: Reference allele",
    "Original Input: Alternate allele",
    "Variant Annotation: Gene",
    "Variant Annotation: cDNA change",
    "Variant Annotation: Protein Change",
    "Variant Annotation: RefSeq",
    "VCF: AF",
    "VCF: FAO",
    "VCF: FDP",
    "VCF: HRUN",
    "VCF: Filter",
    "VCF: Genotype",
    "COSMIC: ID",
    "COSMIC: Variant Count",
    "COSMIC: Variant Count (Tissue)",
    "ClinVar: ClinVar ID",
    "ClinVar: Clinical Significance",
    "gnomAD3: Global AF",
    "PhyloP: Vert Score",
    "CADD: Phred",
    "PolyPhen-2: HDIV Prediction",
    "SIFT: Prediction",
    "VCF: FSAF",
    "VCF: FSAR",
    "VCF: FSRF",
    "VCF: FSRR",
    "VCF: Fisher Odds Ratio",
    "VCF: Fisher P Value",
    "VCF: Binom Proportion",
    "VCF: Binom P Value",
    "Mpileup Qual: Read Depth",
    "Mpileup Qual: Start Reads",
    "Mpileup Qual: Stop Reads",
    "Mpileup Qual: Filtered Reference Forward Read Depth",
    "Mpileup Qual: Filtered Reference Reverse Read Depth",
    "Mpileup Qual: Unfiltered Reference Forward Read Depth",
    "Mpileup Qual: Unfiltered Reference Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Forward Read Depth",
    "Mpileup Qual: Filtered Variant Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Binomial Proportion",
    "Mpileup Qual: Filtered Variant Binomial P Value",
    "Mpileup Qual: Filtered Variant Fishers Odds Ratio",
    "Mpileup Qual: Filtered Variant Fishers P Value",
    "Mpileup Qual: Filtered VAF",
    "Mpileup Qual: Unfiltered Variant Forward Read Depth",
    "Mpileup Qual: Unfiltered Variant Reverse Read Depth",
    "Mpileup Qual: Unfiltered Variant Binomial Proportion",
    "Mpileup Qual: Unfiltered Variant Binomial P Value",
    "Mpileup Qual: Unfiltered Variant Fishers Odds Ratio",
    "Mpileup Qual: Unfiltered Variant Fishers P Value",
    "Mpileup Qual: Unfiltered VAF",
    "VCF: LEN",
    "VCF: QD",
    "VCF: STB",
    "VCF: STBP",
    "VCF: SVTYPE",
    "VCF: TYPE",
    "VCF: QUAL",
    "Variant Annotation: Coding",
    "Variant Annotation: Sequence Ontology",
    "Variant Annotation: Transcript",
    "Variant Annotation: All Mappings",
    "UniProt (GENE): Accession Number",
    "dbSNP: rsID",
    "MDL: Sample Count",
    "MDL: Variant Frequency",
    "MDL: Sample List",
    "amp_ID",
    "Cytoband",
    "MANE_transcript (GRCh38)",
    "Genexus_transcript (GRCh37)",
    "Genexus_Exon(s)",
    "Genexus_codons",
    "tier",
    "test_tissue",
    "Disposition",
]

# CLASSES ------------------------------------------------


class DataModel:
    def __init__(self) -> None:
        self.variables = dict()
        self.filename = str()
        self.selected_disposistion = "None"
        self.selection_index = 0

        self.bdb = BedLookupTable()
        self.idb = InfotrackLookupTable()

        self.DATA_FIELDS = list()

        return None

    def determine_data_fields(self, excel_columns) -> None:
        """Method to determine if this is an already annotated excel file, or if it is an original one.  Sets the data field (columns) list appropriately."""

        return None

    def load_file(self) -> None:
        """Recieves the filename from the view, attempts to load."""

        workbook = openpyxl.load_workbook(self.filename, data_only=True, read_only=True)
        self.variant_list = list()

        # Read in the columns names and make a list, then check the length \
        # to see if it's been opened before.
        self.column_list = [cell.value for cell in workbook.worksheets[0][1]]
        if len(self.column_list) < 70:
            self.new_file = True
        else:
            self.new_file = False

        # capturing the workbook sheet names as dispositions
        for disposition in workbook.sheetnames:
            # Grab the appropriate worksheet and make it active
            sheet = workbook[disposition]

            for row in sheet.iter_rows(min_row=2):
                # prep a dictionary to hold row data
                row_dict = dict()

                # Step through all data fields
                for x in range(len(self.column_list)):
                    try:
                        row_dict[self.column_list[x]] = row[x].value
                    except:
                        row_dict[self.column_list[x]] = None

                # making some temp variables for clarity during coding
                gene = row_dict["Variant Annotation: Gene"]
                bp = int(row_dict["Original Input: Pos"])
                c_dot = row_dict["Variant Annotation: cDNA change"]

                # self.column_list = [*self.column_list, *ADDON_LIST]

                if self.new_file:
                    # ANNOTATING FROM THE BED FILE ---------------------------
                    for item in ADDON_LIST[:-2]:
                        row_dict[item] = self.bdb.get_data(gene, bp, item)
                    # ANNOTATING FROM THE INFOTRACK DB FILE ------------------
                    row_dict["tier"] = self.idb.recommend_tier(
                        gene, c_dot, "Blood"
                    )  # Genexys is just blood samples currently
                    row_dict["test_tissue"] = self.idb.get_tissue_list(gene, c_dot)

                else:
                    extended_range = (
                        len(self.column_list),
                        len(self.column_list) + len(ADDON_LIST),
                    )
                    for x in range(extended_range[0], extended_range[1]):
                        try:
                            row_dict[self.column_list[x]] = row[x].value
                        except:
                            row_dict[self.column_list[x]] = None

                # ignores any sheet names that "non-standard" dispositions.
                # The first sheet in the excel file typically has such a name.
                if disposition in DISPOSITIONS:
                    row_dict["Disposition"] = disposition
                else:
                    row_dict["Disposition"] = "None"

                self.variant_list.append(row_dict.copy())

        return None

    def save_file(self) -> None:
        """ Writes the sorted data out to disk with marker on the filename to \
            denote file has been processed. """

        # appending a "(sorted)" note to the filename, but checking if that \
        # note is already there
        if VCF_FILE_SETTINGS["filename_addon"] in self.filename:
            filename = self.filename
        else:
            filename = (
                self.filename[: -(len(VCF_FILE_SETTINGS["excel_extension"]))]
                + VCF_FILE_SETTINGS["filename_addon"]
                + VCF_FILE_SETTINGS["excel_extension"]
            )

        # Openpyxl package work
        wb = openpyxl.Workbook()
        for tab in DISPOSITIONS:
            wb.create_sheet(tab)
            wb.active = wb[tab]
            ws = wb.active
            row = [column for column in self.column_list]
            ws.append(row)
            for entry in self.variant_list:
                if entry["Disposition"] == tab:
                    row = [entry[x] for x in self.column_list]
                    ws.append(row)

        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))  # removing the default \
        # "Sheet" from openpyxl
        wb.save(filename)

        return None

    def change_disposition(
        self, selection: str, disposition: str, update_dict: dict
    ) -> None:
        """Updates the disposition of the selected record."""

        # Updating all the fields in the model instance of the record
        for field in DATA_FIELDS:
            self.variant_list[selection][field] = update_dict[field]
        self.variant_list[selection]["Disposition"] = disposition

        return None

    def count_dispositions(self, disposition) -> int:
        """ Method to count the number of records in the variant list with \
            the passed disposition. """

        counter = 0

        for row in self.variant_list:
            if disposition == row["Disposition"]:
                counter += 1

        return counter

    def output_text_files(self, *args) -> None:
        """ Placeholder method which will eventually output the needed text \
            files for the reporting script. """

        # pull out the MD number from the filename
        MD_number = (os.path.split(self.filename)[1]).split(".")[0]
        file_location = os.path.split(self.filename)[0]

        # creating filenames, note that the "XXX" replaces the build version \
        # number
        filenames = [
            os.path.join(file_location, (MD_number + x))
            for x in ["_low_coverage.tsv", "_mutations.tsv", "_vus.tsv"]
        ]

        # open all three files at once with a context manager
        with open(filenames[0], "w", encoding="ascii") as file_low, open(
            filenames[1], "w", encoding="ascii"
        ) as file_mut, open(filenames[2], "w", encoding="ascii") as file_vus:
            # write the headers for the tsvs
            file_low.write("Gene	Amplicon	Exon	Codon	Depth\n")
            file_mut.write(
                "Gene	DNA	Protein	VAF^1	COSMIC^2	Tier^3	\
                           Cytoband\n"
            )
            file_vus.write(
                "Gene	DNA	Protein	VAF^1	COSMIC^2	Tier^3	\
                           Cytoband\n"
            )

            # now step through the variant list and write each line out to \
            # the text files as necessary
            for variant in self.variant_list:
                if variant["Disposition"] == "Low VAF Variants":
                    text_string = "{}\t{}\t{}\t{}\t{}\n".format(
                        variant["Variant Annotation: Gene"],
                        variant["amp_ID"],
                        variant["Genexus_Exon(s)"],
                        variant["Genexus_codons"],
                        variant["VCF: FDP"],
                    )
                    file_low.write(text_string)
                elif variant["Disposition"] == "VUS":
                    text_string = "{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(
                        variant["Variant Annotation: Gene"],
                        variant["Variant Annotation: cDNA change"],
                        variant["Variant Annotation: Protein Change"],
                        variant["VCF: AF"],
                        variant["COSMIC: ID"],
                        variant["tier"],
                        variant["Cytoband"],
                    )
                    file_vus.write(text_string)
                elif variant["Disposition"] == "Oncogenic":
                    text_string = "{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(
                        variant["Variant Annotation: Gene"],
                        variant["Variant Annotation: cDNA change"],
                        variant["Variant Annotation: Protein Change"],
                        variant["VCF: AF"],
                        variant["COSMIC: ID"],
                        variant["tier"],
                        variant["Cytoband"],
                    )
                    file_mut.write(text_string)
                else:
                    pass

        message = "TSV text files exported"
        detail = "Text files were successfully written\n" "to the VCF file location.\n"
        mb.showinfo(title="Export Complete", message=message, detail=detail)

        return None


# MAIN LOOP ----------------------------------------------


def main() -> None:
    model = DataModel()

    return None


if __name__ == "__main__":
    main()
