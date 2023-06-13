# CSViewer.py
''' An application to view a csv file. '''

# IMPORTS ------------------------------------------------

# import tkinter as tk 
import ttkbootstrap as tk 
from tkinter import filedialog as fd
from tkinter import messagebox as mb
from tkinter import Widget
import openpyxl
import os 
import json

# VARIABLES ----------------------------------------------

settings_filename = 'settings.json'
if os.path.exists(settings_filename):
    SETTINGS = json.load(settings_filename)
else:
    SETTINGS = {
    }

VCF_FIELDS = [
    "Original Input: Chrom",
    "Original Input: Pos",
    "Original Input: Reference allele",
    "Original Input: Alternate allele",
    "Variant Annotation: Gene",
    "Variant Annotation: cDNA change",
    "Variant Annotation: Protein Change",
    "Variant Annotation: RefSeq",
    "VCF: AF",
    "VCF: FAO",
    "VCF: FDP",
    "VCF: HRUN",
    "VCF: Filter",
    "VCF: Genotype",
    "COSMIC: ID",
    "COSMIC: Variant Count",
    "COSMIC: Variant Count (Tissue)",
    "ClinVar: ClinVar ID",
    "ClinVar: Clinical Significance",
    "gnomAD3: Global AF",
    "PhyloP: Vert Score",
    "CADD: Phred",
    "PolyPhen-2: HDIV Prediction",
    "SIFT: Prediction",
    "VCF: FSAF",
    "VCF: FSAR",
    "VCF: FSRF",
    "VCF: FSRR",
    "VCF: Fisher Odds Ratio",
    "VCF: Fisher P Value",
    "VCF: Binom Proportion",
    "VCF: Binom P Value",
    "Mpileup Qual: Read Depth",
    "Mpileup Qual: Start Reads",
    "Mpileup Qual: Stop Reads",
    "Mpileup Qual: Filtered Reference Forward Read Depth",
    "Mpileup Qual: Filtered Reference Reverse Read Depth",
    "Mpileup Qual: Unfiltered Reference Forward Read Depth",
    "Mpileup Qual: Unfiltered Reference Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Forward Read Depth",
    "Mpileup Qual: Filtered Variant Reverse Read Depth",
    "Mpileup Qual: Filtered Variant Binomial Proportion",
    "Mpileup Qual: Filtered Variant Binomial P Value",
    "Mpileup Qual: Filtered Variant Fishers Odds Ratio",
    "Mpileup Qual: Filtered Variant Fishers P Value",
    "Mpileup Qual: Filtered VAF",
    "Mpileup Qual: Unfiltered Variant Forward Read Depth",
    "Mpileup Qual: Unfiltered Variant Reverse Read Depth",
    "Mpileup Qual: Unfiltered Variant Binomial Proportion",
    "Mpileup Qual: Unfiltered Variant Binomial P Value",
    "Mpileup Qual: Unfiltered Variant Fishers Odds Ratio",
    "Mpileup Qual: Unfiltered Variant Fishers P Value",
    "Mpileup Qual: Unfiltered VAF",
    "VCF: LEN",
    "VCF: QD",
    "VCF: STB",
    "VCF: STBP",
    "VCF: SVTYPE",
    "VCF: TYPE",
    "VCF: QUAL",
    "Variant Annotation: Coding",
    "Variant Annotation: Sequence Ontology",
    "Variant Annotation: Transcript",
    "Variant Annotation: All Mappings",
    "UniProt (GENE): Accession Number",
    "dbSNP: rsID",
    "MDL: Sample Count",
    "MDL: Variant Frequency",
    "MDL: Sample List",
    "Disposition",
]

TOOLTIPS = {
    "Disposition":'What to call this variant.',
    "Original Input: Chrom":'Chromosome on which this gene is found.',
    "Original Input: Pos":'Base pair position of the gene on the chromosome.',
    "Original Input: Reference allele":'Expected finding at this base pair location.',
    "Original Input: Alternate allele":'Specimen finding at this base pair location.',
    "Variant Annotation: Gene":'Gene currently selected from the variant list.',
    "Variant Annotation: cDNA change":'Alteration in the DNA at this location.',
    "Variant Annotation: Protein Change":'Resultant alteration in the protein at this location.',
    "VCF: AF":'Allele fraction of reads with this variant.',
    "VCF: FAO":"Variant read depth at this base pair location, reported by Genexys.",
    "VCF: FDP":"Total read depth at this base pair location, reported by Genexys",
    "VCF: HRUN":"Homopolymer run count, reported by Genexys.",
    "VCF: Filter":"Final filter disposition as given by the Genexys.\n Preferred value: 'PASS'",
    "VCF: Genotype":"Genotype distinction made by the Genexys analyzer.\n 1/1 = homozygous, 0/1 = heterozygous, 0/0 = ???, ./. = ???",
    "COSMIC: ID":"COSMIC website ID for this variant.",
    "COSMIC: Variant Count":"Times this variant has been reported to COSMIC.",
    "COSMIC: Variant Count (Tissue)":"JSON-style dictionary breakdown of tissue types reported to COSMIC for this variant.", # Very long text, needs wordwrap
    "ClinVar: ClinVar ID":"ClinVar website ID for this variant.",
    "ClinVar: Clinical Significance":"Clinical significance as reported by ClinVar website.",
    "gnomAD3: Global AF":"Frequency of this variant being found in the human population, as reported by GnomAD website.",
    "PhyloP: Vert Score":"Vertebrate score for gene conservancy, as reported by web resources.",
    "CADD: Phred":"Combine Annotation Dependent Depletion score, rating pathogenicity.\nRange: 0 = Benign, 48 = Pathogenic",
    "PolyPhen-2: HDIV Prediction":"Score assessing the possible change in phenotype of the protein structure, based on AA change.",
    "SIFT: Prediction":"SNP mutagenesis prediction score based on AA change, as reported by SIFT website.",
    "VCF: FSAF":"Forward Variant Read Depth, reported by Genexys\nValid value: > 10",
    "VCF: FSAR":"Reverse Variant Read Depth, reported by Genexys\nValid value: > 10",
    "VCF: FSRF":"Forward Reference Read Depth, reported by Genexys.",
    "VCF: FSRR":"Reverse Reference Read Depth, reported by Genexys.",
    "VCF: Fisher Odds Ratio":"Fisher's odds ratio calculation, based on Genexys read depth data.",
    "VCF: Fisher P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "VCF: Binom Proportion":"Binomial proportion caclucation, based on Genexys read depth data.",
    "VCF: Binom P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "Mpileup Qual: Read Depth":"Total read depth, as reported by M-Pileup data.\nValid: > 500",
    "Mpileup Qual: Start Reads":"Count of read start signals (strand termination), as reported in the M-Pileup data.",
    "Mpileup Qual: Stop Reads":"Cout of read stop signals (strand initiation), as reported in the M-Pileup data.",
    "Mpileup Qual: Filtered Reference Forward Read Depth":"Forward Reference Read Depth, reported by M-Pileup, @Q20\nValid value: > 10",
    "Mpileup Qual: Filtered Reference Reverse Read Depth":"Reverse Reference Read Depth, reported by M-Pileup, @Q20\nValid value: > 10",
    "Mpileup Qual: Unfiltered Reference Forward Read Depth":"Forward Reference Read Depth, reported by M-Pileup, @Q1\nValid value: > 10",
    "Mpileup Qual: Unfiltered Reference Reverse Read Depth":"Reverse Reference Read Depth, reported by M-Pileup, @Q1\nValid value: > 10",
    "Mpileup Qual: Filtered Variant Forward Read Depth":"Forward Variant Read Depth, reported by M-Pileup, @Q20\nValid value: > 10",
    "Mpileup Qual: Filtered Variant Reverse Read Depth":"Reverse Variant Read Depth, reported by M-Pileup, @Q20\nValid value: > 10",
    "Mpileup Qual: Filtered Variant Binomial Proportion":"Binomial proportion caclucation, based on filtered M-Pileup read depth data.",
    "Mpileup Qual: Filtered Variant Binomial P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "Mpileup Qual: Filtered Variant Fishers Odds Ratio":"Fisher's odds ratio calculation, based on filtered M-Pileup read depth data.",
    "Mpileup Qual: Filtered Variant Fishers P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "Mpileup Qual: Unfiltered Variant Forward Read Depth":"Forward Variant Read Depth, reported by M-Pileup, @Q1\nValid value: > 10",
    "Mpileup Qual: Unfiltered Variant Reverse Read Depth":"Reverse Variant Read Depth, reported by M-Pileup, @Q1\nValid value: > 10",
    "Mpileup Qual: Unfiltered Variant Binomial Proportion":"Binomial proportion caclucation, based on unfiltered M-Pileup read depth data.",
    "Mpileup Qual: Unfiltered Variant Binomial P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "Mpileup Qual: Unfiltered Variant Fishers Odds Ratio":"Fisher's odds ratio calculation, based on unfiltered M-Pileup read depth data.",
    "Mpileup Qual: Unfiltered Variant Fishers P Value":"Statistical p-value.\nLess than 0.05 is preferred.",
    "VCF: LEN":"Length of the variant, as reported by Genexys.",
    "VCF: QD":"???",
    "VCF: STB":"Proprietary strand bias calculation, as reported by Genexys.",
    "VCF: STBP":"Statistical p-value.\nLess than 0.05 is preferred.",
    "VCF: SVTYPE":"Unused data field from the Genexys report.",
    "VCF: TYPE":"Type of variant, as reported by Genexys.",
    "VCF: QUAL":"Quality determination tag, as reported by Genexys.",
    "Variant Annotation: Coding":"Reported coding region variant, as reported by Genexys.",
    "Variant Annotation: Sequence Ontology":"Type of variant/mutation encountered.\n Possible Types: MIS, INT, FSI, IND, SYN, SPL ",
    "Variant Annotation: Transcript":"Ensembl transcript designation code.",
    "Variant Annotation: All Mappings":"JSON-style dictionary breakdown of tissue types present in ??? knowledgebase for this variant.", # Very long text, needs wordwrap
    "UniProt (GENE): Accession Number":"UniProt web resource for the affected protein and biological functions.",
    "dbSNP: rsID":"ID number for the free dbSNP web resource listing of this variant.",
    "MDL: Sample Count":"Instance count of samples with this variant in ",
    "MDL: Variant Frequency":"",
    "MDL: Sample List":"JSON-style dictionary breakdown of tissue types present in ??? knowledgebase for this variant.", # Very long text, needs wordwrap
}

DISPOSITIONS = [
    "None",
    "Harmful",
    "VUS",
    "Low VAF Variants",
    "FLT3 ITDs",
    "Hotspot Exceptions",
]

# CLASSES ------------------------------------------------

class DataModel():

    def __init__(self) -> None:
        self.filename = str()
        self.selected = dict()
        for x in VCF_FIELDS:
            self.selected[x] = None
        self.disposition_counts = dict()
        for x in DISPOSITIONS:
            self.disposition_counts[x] = 0
        self.variant_list = dict()
        return
    
    def _select_variant(self, index:int):
        if index == None:
            for x in VCF_FIELDS:
                self.selected[x] = None
        else:
            for x in VCF_FIELDS:
                self.selected[x] = self.variant_list[index][x]
        
    def _load_file(self, filename):
        """ Recieves the filename from the view, attempts to load. """
        self.filename = filename
        if ".XLSX" in str(self.filename).upper():
            try:
                self._read_excel_file_to_dictionary(self.filename)
            except:
                print("File not read correctly.")
        return
        
    def _read_excel_file_to_dictionary(self, filename:str):
        """ Loads an Excel worksheet, then reads all sheets for variant information. """
        workbook = openpyxl.load_workbook(filename, data_only=True, read_only=True)
        self.variant_list = list()
        for disposition in workbook.sheetnames:
            sheet = workbook[disposition]
            for row in sheet.iter_rows(min_row=2):
                row_dict = dict()
                for x in range(len(VCF_FIELDS)):
                    row_dict[VCF_FIELDS[x]] = row[x].value
                row_dict['Disposition'] = disposition
                self.variant_list.append(row_dict.copy())
        return
    
    def _count_dispositions(self):
        for x in DISPOSITIONS:
            self.disposition_counts[x] = 0
            for variant in self.variant_list:
                self.disposition_counts[variant['Disposition']] += 1
        return


class MainMenu(tk.Menu):
    '''The Application's main menu.'''

    def _event(self, sequence):
        def callback(*_):
            root = self.master.winfo_toplevel()
            root.event_generate(sequence)
        return callback

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        # self.settings = settings
        
        # File Menu --------------------------------
        menu_file = tk.Menu(self, tearoff=False)
        menu_file.add_command(
            label="Open", 
            command=self._event('<<FileSelect>>')
        )
        menu_file.add_command(
            label="Clear", 
            command=self._event('<<FileClear>>')
        )
        menu_file.add_command(
            label="Save", 
            command=self._event('<<FileSave>>')
        )
        menu_file.add_separator()
        menu_file.add_command(
            label="Quit",
            command=self._event('<<FileQuit>>')
        )
        
        # Go Menu ----------------------------------
        menu_go = tk.Menu(self, tearoff=False)
        menu_go.add_command(
            label = "Export Text Files",
            command=self._event('<<ExportTextFiles>>')
        )

        # Themes Menu -------------------------------
        menu_theme = tk.Menu(self, tearoff=0)
        menu_theme.add_command(
            label='Cosmo', 
            command=self._event('<<ThemeCosmo>>')
        )
        menu_theme.add_command(
            label='Flatly', 
            command=self._event('<<ThemeFlatly>>')
        )
        menu_theme.add_command(
            label='Journal', 
            command=self._event('<<ThemeJournal>>')
        )
        menu_theme.add_command(
            label='Litera', 
            command=self._event('<<ThemeLitera>>')
        )
        menu_theme.add_command(
            label='Lumen', 
            command=self._event('<<ThemeLumen>>')
        )
        menu_theme.add_command(
            label='Pulse', 
            command=self._event('<<ThemePulse>>')
        )
        menu_theme.add_command(
            label='Sandstone', 
            command=self._event('<<ThemeSandstone>>')
        )
        menu_theme.add_command(
            label='United', 
            command=self._event('<<ThemeUnited>>')
        )
        menu_theme.add_command(
            label='Yeti', 
            command=self._event('<<ThemeYeti>>')
        )
        menu_theme.add_separator()
        menu_theme.add_command(
            label='Superhero', 
            command=self._event('<<ThemeSuperhero>>')
        )
        menu_theme.add_command(
            label='Darkly', 
            command=self._event('<<ThemeDarkly>>')
        )
        menu_theme.add_command(
            label='Cyborg', 
            command=self._event('<<ThemeCyborg>>')
        )

        # Help Menu ----------------------------
        menu_help = tk.Menu(self, tearoff=False)
        menu_help.add_command(
            label='About...',
            command=self.show_about
        )

        # Add menus in the right order...
        self.add_cascade(label='File', menu=menu_file)
        self.add_cascade(label='Go', menu=menu_go)
        self.add_cascade(label='Theme', menu=menu_theme)
        self.add_cascade(label='Help', menu=menu_help)

        return

    def show_about(self):
        '''Show the about dialog'''
        
        about_message = 'VCF Data Viewer'
        about_detail = (
            'Written by Vektra Casler MD \n'
            '\n'
            'For assistance, please contact \n'
            'vektra_casler@urmc.rochester.edu'
        )
        mb.showinfo(
            title='About',
            message=about_message,
            detail=about_detail
        )

class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 57
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(
            tw, 
            text=self.text, 
            justify=tk.LEFT,
            background="#ffffe0", 
            relief=tk.SOLID, 
            borderwidth=1,
            font=("tahoma", "8", "normal")
        )
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class Application(tk.Window):

    def __init__(self) -> None:
        super().__init__()

        self.data_model = DataModel()

        # Root Window
        self.title('VCF Result Viewer')
        self.resizable(True, True)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.geometry('800x600')

        # Menu
        self.menu = MainMenu(self)
        self.config(menu=self.menu)
        event_callbacks={
            '<<FileSelect>>': self.data_model._load_file(fd),
            # '<<FileClear>>': self._on_file_select,
            # '<<FileSave>>': self._on_file_select,
            '<<FileQuit>>': lambda _: self.quit(),
            # '<<ExportTextFiles>>': self._export_text_files,
            '<<ThemeCosmo>>': lambda _: self.style.theme_use('cosmo'),
            '<<ThemeFlatly>>': lambda _: self.style.theme_use('flatly'),
            '<<ThemeJournal>>': lambda _: self.style.theme_use('journal'),
            '<<ThemeLitera>>': lambda _: self.style.theme_use('litera'),
            '<<ThemeLumen>>': lambda _: self.style.theme_use('lumen'),
            '<<ThemePulse>>': lambda _: self.style.theme_use('pulse'),
            '<<ThemeSandstone>>': lambda _: self.style.theme_use('sandstone'),
            '<<ThemeUnited>>': lambda _: self.style.theme_use('united'),
            '<<ThemeYeti>>': lambda _: self.style.theme_use('yeti'),
            '<<ThemeSuperhero>>': lambda _: self.style.theme_use('superhero'),
            '<<ThemeDarkly>>': lambda _: self.style.theme_use('darkly'),
            '<<ThemeCyborg>>': lambda _: self.style.theme_use('cyborg'),
        }
        for sequence, callback in event_callbacks.items():
            self.bind(sequence, callback)

        # Variables

        self.labels = dict()
        self.labels['filename'] = ('No CSV File Loaded')
        for x in VCF_FIELDS:
            self.labels[x] = tk.Label()

        self.vars = dict()
        self.vars['filename'] = tk.StringVar()
        self.vars['status_text'] = tk.StringVar()
        self.vars['Disposition'] = dict()
        for x in DISPOSITIONS:
            self.vars['Disposition'][x] = tk.IntVar()

        self.variant = dict()
        for x in VCF_FIELDS:
            self.variant[x] = tk.StringVar()
            
        self.buttons = dict()
        self.frames = dict()
        self.radio_buttons = dict()
        self.treeviews = dict()
        self.scrollbars = dict()

        self.validation = dict()
        self.validation['p-values'] = [
            "VCF: Binom P Value",
            "VCF: Fisher P Value",
            "Mpileup Qual: Filtered Variant Binomial P Value",
            "Mpileup Qual: Filtered Variant Fishers P Value",
            "Mpileup Qual: Unfiltered Variant Binomial P Value",
            "Mpileup Qual: Unfiltered Variant Fishers P Value",
            "VCF: STBP"
        ]

        # Widgets ------------------------------------------------------------------------

        # Base Frame
        self.frames['base'] = tk.Frame(self)
        self.frames['base'].pack(expand=True, fill='both',ipadx=10, ipady=10)

        # Left Frame
        self.frames['left'] = tk.LabelFrame(self.frames['base'], text="VCF File Info", relief='groove')
        self.frames['left'].pack(side='left', expand=False, fill='y',ipadx=10, ipady=10, padx=5, pady=5)

        #File load button
        self.labels['filename'] = tk.Label(self.frames['left'], textvariable=None, relief='groove')
        self.labels['filename'].pack(side='top', expand=False, fill='x',ipady=5, padx=5, pady=5)
        self.buttons['load_file'] = tk.Button(self.frames['left'], text="Load a File", command=('<<FileSelect>>'))
        self.buttons['load_file'].pack(side='top', expand=False, fill='x', ipady=5, padx=5, pady=5)

        # Treeview Frame
        self.frames['treeview'] = tk.Frame(self.frames['left'])
        self.frames['treeview'].pack(side='top',expand=True,fill='both', padx=5)

        # Treeview list
        self.treeviews['variant_list'] = tk.Treeview(self.frames['treeview'], columns=VCF_FIELDS, displaycolumns=[5,0], selectmode='browse', show='headings')
        for x in VCF_FIELDS:
            self.treeviews['variant_list'].heading(x, text=x, anchor='center')
        self.treeviews['variant_list'].heading(5, text='Variant')
        self.treeviews['variant_list'].heading(0, text='Disposition')
        self.treeviews['variant_list'].column(column=0, width=100, anchor='center')
        self.treeviews['variant_list'].column(column=5, width=100, anchor='center')
        self.treeviews['variant_list'].pack(side='left', expand=True, fill='both')
        # self.treeviews['variant_list'].bind('<<TreeviewSelect>>', self.item_selected)

        # Treeview Scrollbar
        self.scrollbars['variant_list'] = tk.Scrollbar(self.frames['treeview'], orient=tk.VERTICAL, command=self.treeviews['variant_list'].yview)
        self.treeviews['variant_list'].configure(yscroll=self.scrollbars['variant_list'].set)
        self.scrollbars['variant_list'].pack(side='left', expand=False, fill='y')

        # Disposition Frame
        self.frames['disposition'] = tk.LabelFrame(self.frames['left'], text="Variant Disposition", relief='groove')
        self.frames['disposition'].pack(side='top', expand=False, fill='x', padx=5)
        self.frames['dispo_1'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_1'].pack(side='top', expand=False, fill='x')
        self.frames['dispo_2'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_2'].pack(side='top', expand=False, fill='x')
        self.frames['dispo_3'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_3'].pack(side='top', expand=False, fill='x')
        self.frames['dispo_4'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_4'].pack(side='top', expand=False, fill='x')
        self.frames['dispo_5'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_5'].pack(side='top', expand=False, fill='x')
        self.frames['dispo_6'] = tk.Frame(self.frames['disposition'])
        self.frames['dispo_6'].pack(side='top', expand=False, fill='x')

        # Disposition labels
        self.labels["None"] = tk.Label(self.frames['dispo_1'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["None"].pack(side='left', expand=False, fill='y', padx=5, pady=5)
        self.labels["Low VAF Variants"] = tk.Label(self.frames['dispo_2'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["Low VAF Variants"].pack(side='left', expand=False, fill='y', padx=5, pady=5)
        self.labels["VUS"] = tk.Label(self.frames['dispo_3'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["VUS"].pack(side='left', expand=False, fill='y', padx=5, pady=5)
        self.labels["Harmful"] = tk.Label(self.frames['dispo_4'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["Harmful"].pack(side='left', expand=False, fill='y', padx=5, pady=5)
        self.labels["FLT3 ITDs"] = tk.Label(self.frames['dispo_5'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["FLT3 ITDs"].pack(side='left', expand=False, fill='y', padx=5, pady=5)
        self.labels["Hotspot Exceptions"] = tk.Label(self.frames['dispo_6'], textvariable=self.vars['Disposition']['None'], width=5, relief='groove')
        self.labels["Hotspot Exceptions"].pack(side='left', expand=False, fill='y', padx=5, pady=5)

        # Radio buttons for disposition
        self.radio_buttons["None"] = tk.Radiobutton(self.frames['dispo_1'], text="None (Unassigned)", variable=self.variant['Disposition'], value='None')
        self.radio_buttons["None"].pack(side='left', expand=False, fill='both')
        self.radio_buttons["Low VAF"] = tk.Radiobutton(self.frames['dispo_2'], text="Low VAF", variable=self.variant['Disposition'], value='Low VAF')
        self.radio_buttons["Low VAF"].pack(side='left', expand=False, fill='both')
        self.radio_buttons["VUS"] = tk.Radiobutton(self.frames['dispo_3'], text="VUS", variable=self.variant['Disposition'], value='VUS')
        self.radio_buttons["VUS"].pack(side='left', expand=False, fill='both')
        self.radio_buttons["Harmful"] = tk.Radiobutton(self.frames['dispo_4'], text="Harmful", variable=self.variant['Disposition'], value='Harmful')
        self.radio_buttons["Harmful"].pack(side='left', expand=False, fill='both')
        self.radio_buttons["FLT3 ITDs"] = tk.Radiobutton(self.frames['dispo_5'], text="FLT3 ITDs", variable=self.variant['Disposition'], value='FLT3 ITDs')
        self.radio_buttons["FLT3 ITDs"].pack(side='left', expand=False, fill='both')
        self.radio_buttons["Hotspot Exceptions"] = tk.Radiobutton(self.frames['dispo_6'], text="Hotspot Exception", variable=self.variant['Disposition'], value='Hotspot Exception')
        self.radio_buttons["Hotspot Exceptions"].pack(side='left', expand=False, fill='both')
        # self.radio_buttons["None"].select()

        # Process output files button
        self.buttons['save_disposition'] = tk.Button(self.frames['left'], text="Save Disposition", command=None, state='disabled')
        self.buttons['save_disposition'].pack(side='top', expand=False, fill='x', padx=5, pady=5, ipady=5)

        # Process output files button
        # tk.Button(self.frames['left'], text="Create Disposition Lists").pack(side='top', expand=False, fill='x', padx=5, pady=5, ipady=5)

        # Right Frame
        self.frames['right'] = tk.Frame(self.frames['base'])
        self.frames['right'].pack(side='left', expand=True, fill='both',ipadx=10, ipady=10)

        # Basic Info Frame
        self.frames['basic_info'] = tk.LabelFrame(self.frames['right'], text='Locus Info')
        self.frames['basic_info'].pack(side='top', expand=False, fill='x', padx=5, pady=5)
        self.frames['basic_info_gene'] = tk.Frame(self.frames['basic_info'])
        self.frames['basic_info_gene'].pack(side='left',expand=False, fill='both', padx=5, pady=5)
        tk.Label(self.frames['basic_info_gene'], text="Gene").pack(side='top',expand=False,fill='x')
        self.labels["Variant Annotation: Gene"] = tk.Label(self.frames['basic_info_gene'], width=8, textvariable=self.variant["Variant Annotation: Gene"], relief='groove', font=('bold', 24, 'bold'))
        self.labels["Variant Annotation: Gene"].pack(side='top',expand=True,fill='both')
        self.frames['basic_info_chrom'] = tk.Frame(self.frames['basic_info'])
        self.frames['basic_info_chrom'].pack(side='left',expand=False,fill='both', padx=5, pady=5)
        tk.Label(self.frames['basic_info_chrom'], text="Chromosome").pack(side='top',expand=False, fill='x')
        self.labels["Original Input: Chrom"] = tk.Label(self.frames['basic_info_chrom'], width=6, textvariable=self.variant["Original Input: Chrom"], relief='groove')
        self.labels["Original Input: Chrom"].pack(side='top',expand=False, fill='x')
        tk.Label(self.frames['basic_info_chrom'], text="Base Pair").pack(side='top',expand=False, fill='x')
        self.labels["Original Input: Pos"] = tk.Label(self.frames['basic_info_chrom'], width=12, textvariable=self.variant["Original Input: Pos"], relief='groove')
        self.labels["Original Input: Pos"].pack(side='top',expand=False, fill='x')
        self.frames['basic_info_changes'] = tk.Frame(self.frames['basic_info'])
        self.frames['basic_info_changes'].pack(side='left',expand=False,fill='both', padx=5, pady=5)
        tk.Label(self.frames['basic_info_changes'], text="DNA Change (c-dot)").pack(side='top',expand=False, fill='x')
        self.labels["Variant Annotation: cDNA change"] = tk.Label(self.frames['basic_info_changes'], text="C-dot", textvariable=self.variant["Variant Annotation: cDNA change"], relief='groove')
        self.labels["Variant Annotation: cDNA change"].pack(side='top',expand=False, fill='x')
        tk.Label(self.frames['basic_info_changes'], text="Protein Change (p-dot)").pack(side='top',expand=False, fill='x')
        self.labels["Variant Annotation: Protein Change"] = tk.Label(self.frames['basic_info_changes'], text="P-dot", width=24, textvariable=self.variant["Variant Annotation: Protein Change"], relief='groove')
        self.labels["Variant Annotation: Protein Change"].pack(side='top',expand=False, fill='x')
        self.frames['basic_info_alleles'] = tk.Frame(self.frames['basic_info'])
        self.frames['basic_info_alleles'].pack(side='left',expand=True,fill='both', padx=5, pady=5)
        tk.Label(self.frames['basic_info_alleles'], text="Ref Allele", anchor='nw').pack(side='top',expand=False, fill='x')
        self.labels["Original Input: Reference allele"] = tk.Label(self.frames['basic_info_alleles'], anchor='w', textvariable=self.variant["Original Input: Reference allele"], relief='groove')
        self.labels["Original Input: Reference allele"].pack(side='top',expand=False, fill='x')
        tk.Label(self.frames['basic_info_alleles'], text="Variant Allele", anchor='w').pack(side='top',expand=False, fill='x')
        self.labels["Original Input: Alternate allele"] = tk.Label(self.frames['basic_info_alleles'], anchor='w', textvariable=self.variant["Original Input: Alternate allele"], relief='groove')
        self.labels["Original Input: Alternate allele"].pack(side='top',expand=False, fill='x')
        self.frames['basic_info_misc'] = tk.Frame(self.frames['basic_info'])
        self.frames['basic_info_misc'].pack(side='left',expand=False,fill='both', padx=5, pady=5)
        tk.Label(self.frames['basic_info_misc'], text="Length of Variant (BP)").pack(side='top',expand=False, fill='x')
        self.labels["VCF: LEN"] = tk.Label(self.frames['basic_info_misc'], textvariable=self.variant["VCF: LEN"], relief='groove')
        self.labels["VCF: LEN"].pack(side='top',expand=True, fill='both')

        # middle frame
        self.frames['middle'] = tk.Frame(self.frames['right'])
        self.frames['middle'].pack(side='top',expand=True,fill='both', padx=5, pady=5)

        # Strand Bias Frame
        self.frames['genexys'] = tk.LabelFrame(self.frames['middle'], text='Strand Bias Data', relief='groove')
        self.frames['genexys'].pack(side='left', expand=False, fill='both', padx=(0,10))
        self.frames['genexys_sb_calc'] = tk.Frame(self.frames['genexys'])
        self.frames['genexys_sb_calc'].pack(side='top', expand=False, fill='both', padx=5, pady=5)
        self.frames['genexys_sb_calc'].rowconfigure(0, weight=1)
        self.frames['genexys_sb_calc'].rowconfigure(1, weight=1)
        self.frames['genexys_sb_calc'].columnconfigure(0, weight=0)
        self.frames['genexys_sb_calc'].columnconfigure(1, weight=5)
        self.frames['genexys_sb_calc'].columnconfigure(2, weight=0)
        self.frames['genexys_sb_calc'].columnconfigure(3, weight=1)
        tk.Label(self.frames['genexys_sb_calc'], text="SB (reported)", anchor='e').grid(column=0, row=0, sticky='news', padx=5)
        self.labels["VCF: STB"] = tk.Label(self.frames['genexys_sb_calc'], textvariable=self.variant["VCF: STB"], relief='groove')
        self.labels["VCF: STB"].grid(column=1, row=0, sticky='news')
        tk.Label(self.frames['genexys_sb_calc'], text="p.", anchor='e').grid(column=2, row=0, sticky='news', padx=5)
        self.labels["VCF: STBP"] = tk.Label(self.frames['genexys_sb_calc'], textvariable=self.variant["VCF: STBP"], relief='groove', width=5)
        self.labels["VCF: STBP"].grid(column=3, row=0, sticky='news')

        # Genexys strand Bias Area
        self.frames['sb_GX'] = tk.LabelFrame(self.frames['genexys'], text="Genexys (calculated)")
        self.frames['sb_GX'].pack(side='top', expand=True, fill='both', padx=5, pady=5)
        for x in range(3):
            self.frames['sb_GX'].rowconfigure(x, weight=1)
        for x in range(2,4):
            self.frames['sb_GX'].columnconfigure(x, weight=1)
        tk.Label(self.frames['sb_GX'], text='Genexys', width=8).grid(column=0, row=0, rowspan=3, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['sb_GX'], text='Fwd').grid(column=2, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_GX'], text='Rev').grid(column=3, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_GX'], text='Ref', anchor='e').grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_GX'], text='Var', anchor='e').grid(column=1, row=2, sticky='news', pady=5, padx=5)
        self.labels['VCF: FSRF'] = tk.Label(self.frames['sb_GX'], textvariable=self.variant["VCF: FSRF"], relief='groove')
        self.labels['VCF: FSRF'].grid(column=2, row=1, sticky='news', pady=5, padx=5)
        self.labels['VCF: FSRR'] = tk.Label(self.frames['sb_GX'], textvariable=self.variant["VCF: FSRR"], relief='groove')
        self.labels['VCF: FSRR'].grid(column=3, row=1, sticky='news', pady=5, padx=5)
        self.labels['VCF: FSAF'] = tk.Label(self.frames['sb_GX'], textvariable=self.variant["VCF: FSAF"], relief='groove')
        self.labels['VCF: FSAF'].grid(column=2, row=2, sticky='news', pady=5, padx=5)
        self.labels['VCF: FSAR'] = tk.Label(self.frames['sb_GX'], textvariable=self.variant["VCF: FSAR"], relief='groove')
        self.labels['VCF: FSAR'].grid(column=3, row=2, sticky='news', pady=5, padx=5)

        # Separator
        tk.Separator(self.frames['sb_GX'], orient='horizontal').grid(column=0, row=4, columnspan=4, sticky='ew', pady=5)
        self.frames['sb_GX_results'] = tk.Frame(self.frames['sb_GX'])
        self.frames['sb_GX_results'].grid(column=0, row=5, columnspan=4, sticky='news')
        self.frames['sb_GX_results'].rowconfigure(0, weight=1)
        self.frames['sb_GX_results'].rowconfigure(1, weight=1)
        self.frames['sb_GX_results'].columnconfigure(0, weight=0)
        self.frames['sb_GX_results'].columnconfigure(1, weight=5)
        self.frames['sb_GX_results'].columnconfigure(2, weight=0)
        self.frames['sb_GX_results'].columnconfigure(3, weight=1)

        # Genexys Stats Area
        tk.Label(self.frames['sb_GX_results'], text="Binom. Prop.", anchor='e').grid(column=0, row=0, sticky='news', padx=5)
        self.labels["VCF: Binom Proportion"] = tk.Label(self.frames['sb_GX_results'], textvariable=self.variant["VCF: Binom Proportion"], relief='groove', anchor='center')
        self.labels["VCF: Binom Proportion"].grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_GX_results'], text="p.", anchor='e').grid(column=2, row=1, sticky='news')
        self.labels["VCF: Binom P Value"] = tk.Label(self.frames['sb_GX_results'], textvariable=self.variant["VCF: Binom P Value"], relief='groove', anchor='center', width=5)
        self.labels["VCF: Binom P Value"].grid(column=3, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_GX_results'], text="Fishers OR", anchor='e').grid(column=0, row=1, sticky='news', padx=5)
        self.labels["VCF: Fisher Odds Ratio"] = tk.Label(self.frames['sb_GX_results'], textvariable=self.variant["VCF: Fisher Odds Ratio"], relief='groove', anchor='center')
        self.labels["VCF: Fisher Odds Ratio"].grid(column=1, row=0, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_GX_results'], text="p.", anchor='e').grid(column=2, row=0, sticky='news')
        self.labels["VCF: Fisher P Value"] = tk.Label(self.frames['sb_GX_results'], textvariable=self.variant["VCF: Fisher P Value"], relief='groove', anchor='center', width=5)
        self.labels["VCF: Fisher P Value"].grid(column=3, row=0, sticky='news', pady=5, padx=5)

        # Q20 Read Bias Area
        self.frames['sb_Q_20'] = tk.LabelFrame(self.frames['genexys'], text="Filtered M-Pileup (calculated)")
        self.frames['sb_Q_20'].pack(side='top', expand=True, fill='both', padx=5, pady=5)
        for x in range(3):
            self.frames['sb_Q_20'].rowconfigure(x, weight=1)
        for x in range(2,4):
            self.frames['sb_Q_20'].columnconfigure(x, weight=1)
        tk.Label(self.frames['sb_Q_20'], text='Q20', width=8).grid(column=0, row=0, rowspan=3, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['sb_Q_20'], text='Fwd').grid(column=2, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_Q_20'], text='Rev').grid(column=3, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_Q_20'], text='Ref', anchor='e').grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_20'], text='Var', anchor='e').grid(column=1, row=2, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Filtered Variant Forward Read Depth"] = tk.Label(self.frames['sb_Q_20'], textvariable=self.variant["Mpileup Qual: Filtered Variant Forward Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Filtered Variant Forward Read Depth"].grid(column=2, row=1, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Filtered Variant Reverse Read Depth"] = tk.Label(self.frames['sb_Q_20'], textvariable=self.variant["Mpileup Qual: Filtered Variant Reverse Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Filtered Variant Reverse Read Depth"].grid(column=3, row=1, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Filtered Reference Forward Read Depth"] = tk.Label(self.frames['sb_Q_20'], textvariable=self.variant["Mpileup Qual: Filtered Reference Forward Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Filtered Reference Forward Read Depth"].grid(column=2, row=2, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Filtered Reference Reverse Read Depth"] = tk.Label(self.frames['sb_Q_20'], textvariable=self.variant["Mpileup Qual: Filtered Reference Reverse Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Filtered Reference Reverse Read Depth"].grid(column=3, row=2, sticky='news', pady=5, padx=5)

        # Separator
        tk.Separator(self.frames['sb_Q_20'], orient='horizontal').grid(column=0, row=4, columnspan=4, sticky='ew', pady=5)
        self.frames['sb_Q_20_results'] = tk.Frame(self.frames['sb_Q_20'])
        self.frames['sb_Q_20_results'].grid(column=0, row=5, columnspan=4, sticky='news')
        self.frames['sb_Q_20_results'].rowconfigure(0, weight=1)
        self.frames['sb_Q_20_results'].rowconfigure(1, weight=1)
        self.frames['sb_Q_20_results'].columnconfigure(0, weight=0)
        self.frames['sb_Q_20_results'].columnconfigure(1, weight=5)
        self.frames['sb_Q_20_results'].columnconfigure(2, weight=0)
        self.frames['sb_Q_20_results'].columnconfigure(3, weight=1)

        # Q20 Stats Area
        tk.Label(self.frames['sb_Q_20_results'], text="Binom. Prop.", anchor='e').grid(column=0, row=0, sticky='news', padx=5)
        self.labels["Mpileup Qual: Filtered Variant Binomial Proportion"] = tk.Label(self.frames['sb_Q_20_results'], textvariable=self.variant["Mpileup Qual: Filtered Variant Binomial Proportion"], relief='groove', anchor='center')
        self.labels["Mpileup Qual: Filtered Variant Binomial Proportion"].grid(column=1, row=0, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_20_results'], text="p.", anchor='e').grid(column=2, row=0, sticky='news')
        self.labels["Mpileup Qual: Filtered Variant Binomial P Value"] = tk.Label(self.frames['sb_Q_20_results'], textvariable=self.variant["Mpileup Qual: Filtered Variant Binomial P Value"], relief='groove', anchor='center', width=5)
        self.labels["Mpileup Qual: Filtered Variant Binomial P Value"].grid(column=3, row=0, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_20_results'], text="Fishers OR", anchor='e').grid(column=0, row=1, sticky='news', padx=5)
        self.labels["Mpileup Qual: Filtered Variant Fishers Odds Ratio"] = tk.Label(self.frames['sb_Q_20_results'], textvariable=self.variant["Mpileup Qual: Filtered Variant Fishers Odds Ratio"], relief='groove', anchor='center')
        self.labels["Mpileup Qual: Filtered Variant Fishers Odds Ratio"].grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_20_results'], text="p.", anchor='e').grid(column=2, row=1, sticky='news')
        self.labels["Mpileup Qual: Filtered Variant Fishers P Value"] = tk.Label(self.frames['sb_Q_20_results'], textvariable=self.variant["Mpileup Qual: Filtered Variant Fishers P Value"], relief='groove', anchor='center', width=5)
        self.labels["Mpileup Qual: Filtered Variant Fishers P Value"].grid(column=3, row=1, sticky='news', pady=5, padx=5)

        # Q1 Read Bias Area
        self.frames['sb_Q_1'] = tk.LabelFrame(self.frames['genexys'], text="Unfiltered M-Pileup (calculated)")
        self.frames['sb_Q_1'].pack(side='top', expand=True, fill='both', padx=5, pady=5)
        for x in range(3):
            self.frames['sb_Q_1'].rowconfigure(x, weight=1)
        for x in range(2,4):
            self.frames['sb_Q_1'].columnconfigure(x, weight=1)
        tk.Label(self.frames['sb_Q_1'], text='Q1', width=8).grid(column=0, row=0, rowspan=3, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['sb_Q_1'], text='Fwd').grid(column=2, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_Q_1'], text='Rev').grid(column=3, row=0, sticky='news', pady=(5,0), padx=5)
        tk.Label(self.frames['sb_Q_1'], text='Ref', anchor='e').grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_1'], text='Var', anchor='e').grid(column=1, row=2, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Unfiltered Reference Forward Read Depth"] = tk.Label(self.frames['sb_Q_1'], width=5, textvariable=self.variant["Mpileup Qual: Unfiltered Reference Forward Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Unfiltered Reference Forward Read Depth"].grid(column=2, row=2, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Unfiltered Reference Reverse Read Depth"] = tk.Label(self.frames['sb_Q_1'], width=5, textvariable=self.variant["Mpileup Qual: Unfiltered Reference Reverse Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Unfiltered Reference Reverse Read Depth"].grid(column=3, row=2, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Unfiltered Variant Forward Read Depth"] = tk.Label(self.frames['sb_Q_1'], width=5, textvariable=self.variant["Mpileup Qual: Unfiltered Variant Forward Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Unfiltered Variant Forward Read Depth"].grid(column=2, row=1, sticky='news', pady=5, padx=5)
        self.labels["Mpileup Qual: Unfiltered Variant Reverse Read Depth"] = tk.Label(self.frames['sb_Q_1'], width=5, textvariable=self.variant["Mpileup Qual: Unfiltered Variant Reverse Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Unfiltered Variant Reverse Read Depth"].grid(column=3, row=1, sticky='news', pady=5, padx=5)

        # Separator
        tk.Separator(self.frames['sb_Q_1'], orient='horizontal').grid(column=0, row=4, columnspan=4, sticky='ew', pady=5)
        self.frames['sb_Q_1_results'] = tk.Frame(self.frames['sb_Q_1'])
        self.frames['sb_Q_1_results'].grid(column=0, row=5, columnspan=4, sticky='news')
        self.frames['sb_Q_1_results'].rowconfigure(0, weight=1)
        self.frames['sb_Q_1_results'].rowconfigure(1, weight=1)
        self.frames['sb_Q_1_results'].columnconfigure(0, weight=0)
        self.frames['sb_Q_1_results'].columnconfigure(1, weight=5)
        self.frames['sb_Q_1_results'].columnconfigure(2, weight=0)
        self.frames['sb_Q_1_results'].columnconfigure(3, weight=1)

        # Q1 Stats Area
        tk.Label(self.frames['sb_Q_1_results'], text="Binom. Prop.", anchor='e').grid(column=0, row=0, sticky='news', padx=5)
        self.labels["Mpileup Qual: Unfiltered Variant Binomial Proportion"] = tk.Label(self.frames['sb_Q_1_results'], textvariable=self.variant["Mpileup Qual: Unfiltered Variant Binomial Proportion"], relief='groove', anchor='center')
        self.labels["Mpileup Qual: Unfiltered Variant Binomial Proportion"].grid(column=1, row=0, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_1_results'], text="p.", anchor='e').grid(column=2, row=0, sticky='news')
        self.labels["Mpileup Qual: Unfiltered Variant Binomial P Value"] = tk.Label(self.frames['sb_Q_1_results'], textvariable=self.variant["Mpileup Qual: Unfiltered Variant Binomial P Value"], relief='groove', anchor='center', width=5)
        self.labels["Mpileup Qual: Unfiltered Variant Binomial P Value"].grid(column=3, row=0, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_1_results'], text="Fishers OR", anchor='e').grid(column=0, row=1, sticky='news', padx=5)
        self.labels["Mpileup Qual: Unfiltered Variant Fishers Odds Ratio"] = tk.Label(self.frames['sb_Q_1_results'], textvariable=self.variant["Mpileup Qual: Unfiltered Variant Fishers Odds Ratio"], relief='groove', anchor='center')
        self.labels["Mpileup Qual: Unfiltered Variant Fishers Odds Ratio"].grid(column=1, row=1, sticky='news', pady=5, padx=5)
        tk.Label(self.frames['sb_Q_1_results'], text="p.", anchor='e').grid(column=2, row=1, sticky='news')
        self.labels["Mpileup Qual: Unfiltered Variant Fishers P Value"] = tk.Label(self.frames['sb_Q_1_results'], textvariable=self.variant["Mpileup Qual: Unfiltered Variant Fishers P Value"], relief='groove', anchor='center', width=5)
        self.labels["Mpileup Qual: Unfiltered Variant Fishers P Value"].grid(column=3, row=1, sticky='news', pady=5, padx=5)

        # Genexys info frame
        self.frames['gx_info'] = tk.LabelFrame(self.frames['middle'], text='Genexys Information')
        self.frames['gx_info'].pack(side='top', expand=False, fill='both', pady=(0,5))
        for x in range(1,5,2):
            self.frames['gx_info'].rowconfigure(x, weight=1)
        for x in range(5):
            self.frames['gx_info'].columnconfigure(x, weight=1)

        # gx info top area
        tk.Label(self.frames['gx_info'], text="Allele Fraction").grid(row=0, column=0, sticky='news', padx=5)
        self.labels["VCF: AF"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: AF"], relief='groove')
        self.labels["VCF: AF"].grid(row=1, column=0, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="Variant Type").grid(row=0, column=1, sticky='news', padx=5)
        self.labels["VCF: TYPE"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: TYPE"], relief='groove')
        self.labels["VCF: TYPE"].grid(row=1, column=1, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="Genotype").grid(row=0, column=2, sticky='news', padx=5)
        self.labels["VCF: Genotype"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: Genotype"], relief='groove')
        self.labels["VCF: Genotype"].grid(row=1, column=2, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="Filter (Genexys)").grid(row=0, column=3, sticky='news', padx=5)
        self.labels["VCF: Filter"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: Filter"], relief='groove')
        self.labels["VCF: Filter"].grid(row=1, column=3, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="Quality Score").grid(row=0, column=4, sticky='news', padx=5)
        self.labels["VCF: QUAL"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: QUAL"], relief='groove')
        self.labels["VCF: QUAL"].grid(row=1, column=4, sticky='news', padx=5, pady=5)

        # Separator
        tk.Separator(self.frames['gx_info'], orient='horizontal').grid(row=2, column=0, columnspan=5, sticky='news')

        # gx info bottom area
        tk.Label(self.frames['gx_info'], text="FAO").grid(row=3, column=0, sticky='news', padx=5)
        self.labels["VCF: FAO"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: FAO"], relief='groove')
        self.labels["VCF: FAO"].grid(row=4, column=0, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="FDP").grid(row=3, column=1, sticky='news', padx=5)
        self.labels["VCF: FDP"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: FDP"], relief='groove')
        self.labels["VCF: FDP"].grid(row=4, column=1, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="HRUN").grid(row=3, column=2, sticky='news', padx=5)
        self.labels["VCF: HRUN"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: HRUN"], relief='groove')
        self.labels["VCF: HRUN"].grid(row=4, column=2, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="QD").grid(row=3, column=3, sticky='news', padx=5)
        self.labels["VCF: QD"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: QD"], relief='groove')
        self.labels["VCF: QD"].grid(row=4, column=3, sticky='news', padx=5, pady=5)
        tk.Label(self.frames['gx_info'], text="SVTYPE (Unused)").grid(row=3, column=4, sticky='news', padx=5)
        self.labels["VCF: SVTYPE"] = tk.Label(self.frames['gx_info'], textvariable=self.variant["VCF: SVTYPE"], relief='groove')
        self.labels["VCF: SVTYPE"].grid(row=4, column=4, sticky='news', padx=5, pady=5)

        # mpileup info frame
        self.frames['mpl_info'] = tk.LabelFrame(self.frames['middle'], text='M-Pileup Information')
        self.frames['mpl_info'].pack(side='top', expand=False, fill='both', pady=5)
        self.frames['mpl_info_RD'] = tk.Frame(self.frames['mpl_info'])
        self.frames['mpl_info_RD'].pack(side='left', expand=True, fill='both', pady=5, padx=5)
        tk.Label(self.frames['mpl_info_RD'], text="Total Read Depth").pack(side='left', expand=False, fill='both')
        self.labels["Mpileup Qual: Read Depth"] = tk.Label(self.frames['mpl_info_RD'], textvariable=self.variant["Mpileup Qual: Read Depth"], relief='groove')
        self.labels["Mpileup Qual: Read Depth"].pack(side='left', expand=True, fill='both', pady=5, padx=5)
        self.frames['mpl_info_starts'] = tk.Frame(self.frames['mpl_info'])
        self.frames['mpl_info_starts'].pack(side='left', expand=True, fill='both', pady=5, padx=5)
        tk.Label(self.frames['mpl_info_starts'], text="Count: Read Starts").pack(side='left', expand=False, fill='both', pady=5, padx=5)
        self.labels["Mpileup Qual: Start Reads"] = tk.Label(self.frames['mpl_info_starts'], textvariable=self.variant["Mpileup Qual: Start Reads"], relief='groove')
        self.labels["Mpileup Qual: Start Reads"].pack(side='left', expand=True, fill='both', pady=5, padx=5)
        self.frames['mpl_info_ends'] = tk.Frame(self.frames['mpl_info'])
        self.frames['mpl_info_ends'].pack(side='left', expand=True, fill='both', pady=5, padx=5)
        tk.Label(self.frames['mpl_info_ends'], text="Count: Read Ends").pack(side='left', expand=False, fill='both', pady=5, padx=5)
        self.labels["Mpileup Qual: Stop Reads"] = tk.Label(self.frames['mpl_info_ends'], textvariable=self.variant["Mpileup Qual: Stop Reads"], relief='groove')
        self.labels["Mpileup Qual: Stop Reads"].pack(side='left', expand=True, fill='both', pady=5, padx=5)

        # MPL Info Frame
        self.frames['other'] = tk.Frame(self.frames['middle'])
        self.frames['other'].pack(side='top',expand=True, fill='both')
        self.frames['var_annot'] = tk.LabelFrame(self.frames['other'], text='Variant Annotation')
        self.frames['var_annot'].pack(side='left',expand=True, fill='both', padx=(0,5))
        for x in range(3):
            self.frames['var_annot'].columnconfigure(x, weight=1)
        self.frames['var_annot'].rowconfigure(3, weight=99)
        tk.Label(self.frames['var_annot'], text="Coding Region").grid(column=0, row=0, sticky='news', padx=5)
        self.labels["Variant Annotation: Coding"] = tk.Label(self.frames['var_annot'], textvariable=self.variant["Variant Annotation: Coding"], relief='groove')
        self.labels["Variant Annotation: Coding"].grid(column=0, row=1, sticky='news', padx=5)
        tk.Label(self.frames['var_annot'], text="Variant Type (Seq. Ontology)").grid(column=1, row=0, sticky='news', padx=5)
        self.labels["Variant Annotation: Sequence"] = tk.Label(self.frames['var_annot'], textvariable=self.variant["Variant Annotation: Sequence Ontology"], relief='groove')
        self.labels["Variant Annotation: Sequence"].grid(column=1, row=1, sticky='news', padx=5)
        tk.Label(self.frames['var_annot'], text="Transcript").grid(column=2, row=0, sticky='news', padx=5)
        self.labels["Variant Annotation: Transcript"] = tk.Label(self.frames['var_annot'], textvariable=self.variant["Variant Annotation: Transcript"], relief='groove')
        self.labels["Variant Annotation: Transcript"].grid(column=2, row=1, sticky='news', padx=5)
        tk.Label(self.frames['var_annot'], text="All Mappings").grid(column=0, row=2, columnspan=3, sticky='news', padx=5)
        self.labels["Variant Annotation: All Mappings"] = tk.Label(self.frames['var_annot'], textvariable=self.variant["Variant Annotation: All Mappings"], relief='groove', wraplength=500)
        self.labels["Variant Annotation: All Mappings"].grid(column=0, columnspan=3, row=3, sticky='news', padx=5, pady=(0,5))

        # MDL Info area
        self.frames['mdl'] = tk.LabelFrame(self.frames['other'], text='MDL Info')
        self.frames['mdl'].pack(side='left',expand=True, fill='both', padx=(5,0))
        self.frames['mdl'].rowconfigure(3, weight=99)
        self.frames['mdl'].columnconfigure(0, weight=1)
        self.frames['mdl'].columnconfigure(1, weight=1)
        tk.Label(self.frames['mdl'], text="Sample Count").grid(column=0, row=0, sticky='news', padx=5)
        self.labels["MDL: Sample Count"] = tk.Label(self.frames['mdl'], textvariable=self.variant["MDL: Sample Count"], relief='groove')
        self.labels["MDL: Sample Count"].grid(column=0, row=1, sticky='news', padx=5)
        tk.Label(self.frames['mdl'], text="Variant Frequency").grid(column=1, row=0, sticky='news', padx=5)
        self.labels["MDL: Variant Frequency"]  = tk.Label(self.frames['mdl'], textvariable=self.variant["MDL: Variant Frequency"], relief='groove')
        self.labels["MDL: Variant Frequency"].grid(column=1, row=1, sticky='news', padx=5)
        tk.Label(self.frames['mdl'], text="Sample List").grid(column=0, columnspan=2, row=2, sticky='news', padx=5)
        self.labels["MDL: Sample List"] = tk.Label(self.frames['mdl'], textvariable=self.variant["MDL: Sample List"], relief='groove', wraplength=500)
        self.labels["MDL: Sample List"].grid(column=0, columnspan=2, row=3, sticky='news', padx=5, pady=(0,5))

        # Web Resources Stats
        self.frames['bottom'] = tk.LabelFrame(self.frames['right'], text='Database Information')
        self.frames['bottom'].pack(side='bottom', expand=False, fill='both', padx=5, pady=5)
        self.frames['bottom'].columnconfigure(0, weight=1)
        self.frames['bottom'].columnconfigure(1, weight=1)
        self.frames['bottom_l'] = tk.Frame(self.frames['bottom'])
        self.frames['bottom_l'].grid(column=0, row=0, sticky='news')
        self.frames['bottom_2'] = tk.Frame(self.frames['bottom'])
        self.frames['bottom_2'].grid(column=1, row=0, sticky='news')

        # COSMIC Tissue Variant Count Region
        self.labels["COSMIC: Variant Count (Tissue)"] = tk.Label(self.frames['bottom_l'], textvariable=self.variant["COSMIC: Variant Count (Tissue)"], relief='groove', wraplength=600)
        self.labels["COSMIC: Variant Count (Tissue)"].pack(side='left', expand=True, fill='both', padx=5, pady=5)

        # COSMIC Area
        self.frames['web_cosmic'] = tk.LabelFrame(self.frames['bottom_2'], text='COSMIC')
        self.frames['web_cosmic'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_cosmic'], text="ID", anchor='center').pack(side='top', expand=False, fill='x', padx=5)
        self.labels["COSMIC: ID"] = tk.Label(self.frames['web_cosmic'], textvariable=self.variant["COSMIC: ID"], relief='groove', width=12)
        self.labels["COSMIC: ID"].pack(side='top', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_cosmic'], text="Count").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["COSMIC: Variant Count"] = tk.Label(self.frames['web_cosmic'], textvariable=self.variant["COSMIC: Variant Count"], relief='groove')
        self.labels["COSMIC: Variant Count"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # Clinvar Area
        self.frames['web_clinvar'] = tk.LabelFrame(self.frames['bottom_2'], text='ClinVar')
        self.frames['web_clinvar'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_clinvar'], text="ID").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["ClinVar: ClinVar ID"] = tk.Label(self.frames['web_clinvar'], textvariable=self.variant["ClinVar: ClinVar ID"], relief='groove')
        self.labels["ClinVar: ClinVar ID"].pack(side='top', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_clinvar'], text="Significance").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["ClinVar: Clinical Significance"] = tk.Label(self.frames['web_clinvar'], textvariable=self.variant["ClinVar: Clinical Significance"], relief='groove')
        self.labels["ClinVar: Clinical Significance"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # Gnomad Area
        self.frames['web_gnomad'] = tk.LabelFrame(self.frames['bottom_2'], text='GnomAD')
        self.frames['web_gnomad'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_gnomad'], text="Global AF").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["gnomAD3: Global AF"] = tk.Label(self.frames['web_gnomad'], textvariable=self.variant["gnomAD3: Global AF"], relief='groove')
        self.labels["gnomAD3: Global AF"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # CADD Area
        self.frames['web_cadd'] = tk.LabelFrame(self.frames['bottom_2'], text='CADD')
        self.frames['web_cadd'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_cadd'], text="Phred Score").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["CADD: Phred"] = tk.Label(self.frames['web_cadd'], textvariable=self.variant["CADD: Phred"], relief='groove')
        self.labels["CADD: Phred"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # PolyPhen Area
        self.frames['web_polyphen'] = tk.LabelFrame(self.frames['bottom_2'], text='PolyPhen-2')
        self.frames['web_polyphen'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_polyphen'], text="HDIV").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["PolyPhen-2: HDIV Prediction"] = tk.Label(self.frames['web_polyphen'], textvariable=self.variant["PolyPhen-2: HDIV Prediction"], relief='groove')
        self.labels["PolyPhen-2: HDIV Prediction"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # SIFT Area
        self.frames['web_sift'] = tk.LabelFrame(self.frames['bottom_2'], text='SIFT')
        self.frames['web_sift'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_sift'], text="Prediction").pack(side='top', expand=False, fill='x', padx=5)
        self.labels["SIFT: Prediction"] = tk.Label(self.frames['web_sift'], textvariable=self.variant["SIFT: Prediction"], relief='groove')
        self.labels["SIFT: Prediction"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # dbSNP Area
        self.frames['web_dbsnp'] = tk.LabelFrame(self.frames['bottom_2'], text='dbSNP')
        self.frames['web_dbsnp'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_dbsnp'], text="rsID").pack(side='top', expand=False, fill='both', padx=5)
        self.labels["dbSNP: rsID"] = tk.Label(self.frames['web_dbsnp'], textvariable=self.variant["dbSNP: rsID"], relief='groove')
        self.labels["dbSNP: rsID"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # UniProt Area
        self.frames['web_dbsnp'] = tk.LabelFrame(self.frames['bottom_2'], text='UniProt (GENE)')
        self.frames['web_dbsnp'].pack(side='left', expand=True, fill='both', padx=5, pady=5)
        tk.Label(self.frames['web_dbsnp'], text="Accession Number").pack(side='top', expand=False, fill='both', padx=5)
        self.labels["UniProt (GENE): Accession Number"] = tk.Label(self.frames['web_dbsnp'], textvariable=self.variant["UniProt (GENE): Accession Number"], relief='groove')
        self.labels["UniProt (GENE): Accession Number"].pack(side='top', expand=True, fill='both', padx=5, pady=5)

        # Create tooltips for labels
        for key,value in TOOLTIPS.items():
            CreateToolTip(self.labels[key], value)

        return
    
    def load_selected_record(self, variant_dict):
        for x in VCF_FIELDS:
            self.variant[x] = self.model.selected[x]
        return

    # def validate_cells(self):

    #     for x in VCF_FIELDS:
    #         self.labels[x]['bg'] = self.color_enabled
    #         self.labels[x]['fg'] = self.color_normal

    #         if not self.variant[x].get():
    #             self.labels[x]['bg'] = self.color_disabled
    #             continue

    #         if x in self.validation['p-values']:
    #             if float(self.variant[x].get()) > 0.05:
    #                 self.labels[x]['fg'] = self.color_warning

    #     return

    # def item_selected(self, event):

    #     for selected_item in self.treeviews['variant_list'].selection():
    #         item = self.treeviews['variant_list'].item(selected_item)
    #         record = item['values']

    #         for x in range(len(VCF_FIELDS)):
    #             self.variant[VCF_FIELDS[x]].set(record[x])
        
    #     self.vars['Disposition'].set(self.variant['Disposition'].get())
    #     if self.vars['Disposition'].get():
    #         self.buttons['save_disposition']['state'] = 'normal'

    #     self.validate_cells()

    #     return
    
    # def save_disposition(self):
        
    #     selection = self.treeviews['variant_list'].focus()
    #     self.treeviews['variant_list'].set(selection, column='Disposition', value=self.vars['Disposition'].get())

    #     self.count_dispositions()
    
    #     return


# FUNCTIONS ----------------------------------------------

def CreateToolTip(widget: Widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


# MAIN LOOP ----------------------------------------------

def main():

    # Mainloop
    root = Application()
    root.mainloop()    

    return

if __name__ == '__main__':
    main()
